#!/usr/bin/env python
from openpyxl import load_workbook


if __name__ == '__main__':

    book = load_workbook(filename = './nodes.xlsx')
    sheet = book.active
    total_rows = len(sheet['A']) 
    total_cols = len(sheet['1'])

    #print(f"rows: {total_rows} | cols: {total_cols}")
    nodes = list()

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=total_rows, max_col=total_cols):
        new_node = list()
        for cell in row:
            new_node.append(cell.value)
        nodes.append(new_node)
    
    for n in nodes:
        for c in n:
            print(f"{c} | ", end="")
        print("\n")
        


            
