#!/usr/bin/env python
from sys import argv, exit
from openpyxl import load_workbook

class CsvGen:
    def __init__(self, sheets):
        self.sheets = sheets
        self.configs = list()

    def read_sheets(self):
        for sheet in self.sheets:
            
            node_list = list()

            file_name = f"./{sheet}"
            book = load_workbook(filename = file_name)
            s = book.active
            total_rows = len(s['A']) 
            total_cols = len(s['1'])

            for row in s.iter_rows(
                min_row= 2,
                min_col= 1,
                max_row = total_rows,
                max_col= total_cols,
            ):
                new_node = list()
                for cell in row:
                    new_node.append(cell.value)
                node_list.append(new_node)
        
            self.configs.append(node_list)
    
    def generate_csv(self):
        for index, config in enumerate(self.configs):
            try:
                with open(f"white_list_{index}.csv", "w+") as new_wl_file:
                    for node in config:
                        line = f"Windows, {node[1]}, {node[2]}, {node[4]}, <description>, <documentation>\n"
                        new_wl_file.write(line)
            except:
                print("[x] File I/O error...")
            finally:
                new_wl_file.close()

def usage():
    print(f"[!] Usage: {argv[0]} < FILE_1.xlsx [FILE_2.xlsx] ...>")
    print("[!] Please supply at least one valid spread sheet in .xlsx format...")
    exit(0)

def validate():
    if len(argv) < 2:
        usage()
    elif argv[1][-5:] != ".xlsx":
        usage()
    else:
        return list(argv[1:])

def run(sheets):
    gen = CsvGen(sheets)
    gen.read_sheets()
    gen.generate_csv()

if __name__ == '__main__':

    sheets = validate()
    run(sheets)
    